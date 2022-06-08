from flask import Flask, jsonify, request, render_template, send_from_directory
from datetime import datetime
from werkzeug.security import safe_join
from flask_cors import CORS
from threading import Thread
import logging
import pickle
import cv2
import imutils
import numpy
import os
from openpyxl import load_workbook, Workbook

logging.basicConfig(format='%(asctime)s [%(levelname)s] %(filename)s:%(lineno)d | %(message)s',
                    datefmt='%d-%m-%Y %H:%M:%S', level=logging.DEBUG)

day = int(datetime.now().strftime("%d"))
month = str(datetime.now().strftime("%b"))
today_full_date = str(datetime.now().strftime('%d-%b-%Y'))
CONFIDANCE_SCORE = 65
static_path = safe_join(os.path.dirname(__file__), 'static')
app = Flask(__name__,
            static_folder="static/static", template_folder="static", static_url_path="/static")
CORS(app)

teachers_db_file = f"db/teachers.pickle"

if not os.path.exists(teachers_db_file):
    pickle.dump({}, open(teachers_db_file, 'wb'))


@app.route("/")
def index():
    return render_template('index.html')


@app.route("/api/teacher_login", methods=["POST"])
def teacher_login():
    try:
        req_body = request.get_json()
        email = req_body["email"]
        password = req_body["password"]

        teachers_db = pickle.load(open(teachers_db_file, 'rb'))

        if email not in teachers_db:
            return jsonify({
                'error': 'Teacher not found'
            }), 400

        if teachers_db[email]['password'] != password:
            return jsonify({
                'error': 'Invalid password'
            }), 400

        return jsonify(teachers_db[email])

    except Exception as e:
        logging.error(e)
        return jsonify({
            'error': 'Server Error'
        }), 500


@app.route("/api/teacher_add", methods=["POST"])
def teacher_add():
    try:
        req_body = request.get_json()
        name = req_body["name"]
        email = req_body["email"]
        subject = req_body["subject"]
        password = req_body["password"]

        if not os.path.exists(teachers_db_file):
            pickle.dump({}, open(teachers_db_file, 'wb'))

        teachers_db = pickle.load(open(teachers_db_file, 'rb'))

        if email in teachers_db:
            return jsonify({
                'error': 'Teacher already exist with same email'
            }), 400

        teachers_db[email] = {
            'name': name,
            'subject': subject,
            'password': password
        }

        pickle.dump(teachers_db, open(teachers_db_file, 'wb'))

        return jsonify(teachers_db[email])

    except Exception as e:
        logging.error(e)
        return jsonify({
            'error': 'Server Error'
        }), 500


def get_count(mylist):
    times = 0
    for occurrence in mylist:
        if occurrence == 'Present':
            times += 1
    return times


MAX_IMAGES_COUNT = 120


def add_student(sub_data):
    haar_file = 'haarcascade_frontalface_default.xml'
    datasets = 'datasets'
    path = os.path.join(datasets, sub_data)
    if not os.path.isdir(path):
        os.mkdir(path)
    (width, height) = (130, 100)  # defining the size of images
    face_cascade = cv2.CascadeClassifier(haar_file)
    webcam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    count = 0
    while count <= MAX_IMAGES_COUNT:
        (_, im) = webcam.read()
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 4)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(im, f"Wait - {MAX_IMAGES_COUNT - count} Images Remaining", (x + 10, y - 10),
                        cv2.FONT_HERSHEY_PLAIN, 1.4, (0, 255, 0), 2)
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (width, height))
            cv2.imwrite('%s/%s.png' % (path, count), face_resize)
            count += 1
        cv2.imshow('OpenCV', im)
        cv2.waitKey(1)

    webcam.release()
    cv2.destroyAllWindows()
    blank_image = numpy.zeros((90, 512, 3), numpy.uint8)
    blank_image[:] = (0, 124, 255)
    cv2.putText(blank_image, f"Student added Successfully!", (20, 50), cv2.FONT_HERSHEY_PLAIN, 2, (255, 255, 255), 2)
    cv2.imshow('OpenCV', blank_image)

    while True:
        key = cv2.waitKey(5)
        if key == 27:
            cv2.destroyAllWindows()
            break


def start_attn(subject):
    _, STUDENT_LIST, _ = next(os.walk('datasets'), ([], [], []))
    excel_file_path = f"excel/attendance-{subject}.xlsx"
    if not os.path.exists(excel_file_path):
        _wb = Workbook()
        _ws = _wb["Sheet"]
        _ws.cell(column=1, row=1, value="Name")
        for j in range(1, 32):
            _ws.cell(column=j + 1, row=1, value=f'{j}/{str(datetime.now().strftime("%b"))}/22')
        _ws.cell(column=33, row=1, value="Percentage")
        _wb.save(excel_file_path)
        _wb.close()

    db_file_path = f"db/subject-{subject}.pickle"
    if not os.path.exists(db_file_path):
        pickle.dump({}, open(db_file_path, 'wb'))

    _names_database = pickle.load(open(db_file_path, 'rb'))
    new_dict = {}

    for _, itm in enumerate(STUDENT_LIST):
        if str(itm) in _names_database:
            new_dict[itm] = _names_database[itm]
        else:
            arr1 = []
            for i in range(0, 32):  # 32 is percent
                arr1.append("")
            new_dict[itm] = arr1
    _names_database = new_dict
    pickle.dump(_names_database, open(db_file_path, 'wb'))

    haar_file = 'haarcascade_frontalface_default.xml'
    datasets = 'datasets'
    print('Training...')
    (images, labels, names, id) = ([], [], {}, 0)
    for (subdirs, dirs, files) in os.walk(datasets):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(datasets, subdir)
            for filename in os.listdir(subjectpath):
                path = subjectpath + '/' + filename
                label = id
                images.append(cv2.imread(path, 0))
                labels.append(int(label))
            id += 1
    (width, height) = (130, 100)
    (images, labels) = [numpy.array(lis) for lis in [images, labels]]
    model = cv2.face.LBPHFaceRecognizer_create(radius=2, neighbors=10, grid_x=8, grid_y=8)
    model.train(images, labels)
    face_cascade = cv2.CascadeClassifier(haar_file)
    webcam = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    student_chk = []
    present_students = []
    last_present_str = ""
    while True:
        _, im = webcam.read()
        im = imutils.resize(im, width=400)
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        cv2.putText(im, f"Confirm Present: {last_present_str}", (20, 20), cv2.FONT_HERSHEY_PLAIN, 1.3, (0, 0, 255), 2)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            face = gray[y:y + h, x:x + w]
            face_resize = cv2.resize(face, (width, height))
            prediction = model.predict(face_resize)
            cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)

            if prediction[1] >= CONFIDANCE_SCORE:
                nam = names[prediction[0]]
                if nam not in student_chk:
                    student_chk.append(nam)
                    present_students.append(nam)
                    STUDENT_LIST.remove(nam)
                    last_present_str = nam
                    _names_database[nam][day - 1] = 'Present'
                    print(f"{nam} - Present on {today_full_date}")

                cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 3)
                cv2.putText(im, '%s - %.0f' % (names[prediction[0]], prediction[1]), (x - 10, y - 10),
                            cv2.FONT_HERSHEY_PLAIN, 1.4, (0, 255, 0), 2)


            else:
                cv2.rectangle(im, (x, y), (x + w, y + h), (0, 0, 255), 3)
                cv2.putText(im, 'not recognized', (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

        cv2.imshow('OpenCV', im)
        key = cv2.waitKey(10)
        if key == 27:
            webcam.release()
            cv2.destroyAllWindows()

            for index, item in enumerate(STUDENT_LIST):
                _names_database[item][day - 1] = 'Absent'

            mywb = load_workbook(filename=excel_file_path)
            myws = mywb["Sheet"]
            myws.cell(column=1, row=1, value="Name")
            for cnt, item in enumerate(_names_database):
                myws.cell(column=1, row=cnt + 2, value=item)  # name
                for _i in range(0, 31):
                    myws.cell(column=_i + 2, row=cnt + 2, value=_names_database[item][_i])
                present_count = get_count(_names_database[item])
                myws.cell(column=33, row=cnt + 2, value=str(round((present_count / 31) * 100, 2)) + ' %')
            mywb.save(filename=excel_file_path)
            mywb.close()
            pickle.dump(_names_database, open(db_file_path, 'wb'))
            break


@app.route("/api/take_attendance", methods=["GET"])
def take_attendance():
    try:
        subject = request.args.get('subject')
        Thread(target=start_attn, args=(subject,), daemon=True).start()
        return jsonify({
            'msg': 'ok'
        }), 200
    except Exception as e:
        logging.error(e)
        return jsonify({
            'error': 'Server Error'
        }), 500


@app.route("/api/add_student", methods=["GET"])
def add_student_req():
    try:
        subject = request.args.get('data')
        Thread(target=add_student, args=(subject,), daemon=True).start()
        return jsonify({
            'msg': 'ok'
        }), 200
    except Exception as e:
        logging.error(e)
        return jsonify({
            'error': 'Server Error'
        }), 500


@app.route("/api/view_stud", methods=["GET"])
def view_stud():
    try:
        subject = request.args.get('subj')
        db_file_path = f"db/subject-{subject}.pickle"

        if not os.path.exists(db_file_path):
            return jsonify({
                'error': 'err'
            }), 500

        _names_database = pickle.load(open(db_file_path, 'rb'))
        return jsonify({
            'data': _names_database
        }), 200
    except Exception as e:
        logging.error(e)
        return jsonify({
            'error': 'Server Error'
        }), 500


@app.route('/excel/<path:path>')
def send_report(path):
    return send_from_directory('excel', path)


@app.route('/<path:path>')
def _static(path):
    if os.path.isdir(safe_join(static_path, path)):
        path = os.path.join(path, 'index.html')
    return send_from_directory(static_path, path)


logging.info(f'Starting Flask Web Server on :8080')
app.run(port=8080, debug=True)
